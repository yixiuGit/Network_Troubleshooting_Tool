import os
import json
import re
import shutil
from deepdiff import DeepDiff
from pprint import pprint
from openpyxl import load_workbook, Workbook
from pyats.contrib.creators.file import File
'''This one will try to collect all agency and site info for the devices
   that have changed, save them in dictionary.
   when create testbed files, delete existing testbed files first, then
   create new. This is because Cisco pyats create can only create testbed
   file if there is none, if testbed file exist with same file name, it will
   throw error.'''
# load the initial json file from spectrum
with open('test.json') as f:
    data = json.load(f)
# load the updated json file
with open('test1.json') as f:
    data1 = json.load(f)

# get difference between 2 files
# ignore_order = true will eliminate the difference only related to position
difference = DeepDiff(data,data1, ignore_order=True)
pprint(difference)

# dictionary that will hold agencies and site that have changed, which will used to delete existing testbed file
dic = {}
# list that will hold agencies that appeared in difference
agencylist = []

# define the template file lcoation
origin_file = r"/home/cxu/pyats/hostname.xlsx"

# the difference will be categorized in 3 groups
#   iterable_item_added
#   iterable_item_removed
#   values_changed

try:
    diffData = difference['iterable_item_added']
    for item in diffData.values():
        agency = item['device_network']
        agencylist.append(agency)
        orgi_sitename = item['device_site']
        temp_sitename = orgi_sitename.replace(":", ' ')
        if agency == 'tnsw_network':
            sitename = item['device_hostname'][4:12]
        # convert sitename that only contains alphnum and space
        else:
            sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
        directory = r"/home/cxu/pyats/{}".format(agency + '/' + sitename)
        # if directory does not exist - new site, create new directory, copy template file to newly created directory
        # append new info to the new file
        if not os.path.exists(directory):
            os.makedirs(directory)
            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
            shutil.copyfile(origin_file, target_file)
            wb = load_workbook(target_file)
            page = wb.active
            info = [item['device_hostname'], item['device_native_ip'], 'c', '', 'ssh', item['device_type']]
            page.append(info)
            wb.save(filename=target_file)
            # if agency is not in agencylist, add new item(agency:sitename) in dic
            # if sitename is not in dic[agency], add sitename as new value to dic[agency]
            # this is to avoid to add the same sitename multiple times to the same dic[agency]
            if agency not in dic.keys():
                dic[agency] = []
                dic[agency].append(sitename)
            elif sitename not in dic[agency]:
                dic[agency].append(sitename)
        else:
            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
            wb = load_workbook(target_file)
            source = wb['Sheet1']
            page = wb.active
            info = [item['device_hostname'], item['device_native_ip'], 'c', '', 'ssh', item['device_type']]
            hostlist = []
            # get all entries in column A of hostname.xlsx, which is device hostname
            for cell in source['A']:
                hostlist.append(cell.value)
            if item['device_hostname'] in hostlist:
                pass
            else:
                page.append(info)
                wb.save(filename=target_file)
                if agency not in dic.keys():
                    dic[agency] = []
                    dic[agency].append(sitename)
                elif sitename not in dic[agency]:
                    dic[agency].append(sitename)
except KeyError:
    print('no new')
try:
    diffData = difference['iterable_item_removed']
    for item in diffData.values():
        agency = item['device_network']
        orgi_sitename = item['device_site']
        temp_sitename = orgi_sitename.replace(":", ' ')
        if agency == 'tnsw_network':
            sitename = item['device_hostname'][4:12]
        else:
            sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
        directory = r"/home/cxu/pyats/{}".format(agency + '/' + sitename)
        if not os.path.exists(directory):
            os.makedirs(directory)
            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
            shutil.copyfile(origin_file, target_file)
            wb = load_workbook(target_file)
            page = wb.active
            info = [item['device_hostname'], item['device_native_ip'], 'c', '', 'ssh', item['device_type']]
            page.append(info)
            wb.save(filename=target_file)
            if agency not in dic.keys():
                dic[agency] = []
                dic[agency].append(sitename)
            elif sitename not in dic[agency]:
                dic[agency].append(sitename)
        else:
            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
            wb = load_workbook(target_file)
            source = wb['Sheet1']
            page = wb.active
            info = [item['device_hostname'], item['device_native_ip'], 'c', '', 'ssh', item['device_type']]
            hostlist = []
            for cell in source['A']:
                hostlist.append(cell.value)
            if item['device_hostname'] in hostlist:
                index = hostlist.index(item['device_hostname'])+1
                source.delete_rows(index)
                wb.save(filename=target_file)
            else:
                page.append(info)
                wb.save(filename=target_file)
                if agency not in dic.keys():
                    dic[agency] = []
                    dic[agency].append(sitename)
                elif sitename not in dic[agency]:
                    dic[agency].append(sitename)
except KeyError:
    print('no remove')

# When have values_changed, consider 2 scenarios,
#   hostname changed, ip may changed
#   ip changed but hostname unchanged
'''{'values_changed': {"root[0]['device_hostname']": {'new_value': 'G1CTLSWI1133',
                                                   'old_value': 'RLC_7347_RT01'},
                    "root[0]['device_model_type_name']": {'new_value': 'CiscoNXOS',
                                                          'old_value': 'GnSNMPDev'},
                    "root[0]['device_native_ip']": {'new_value': '10.33.144.62',
                                                    'old_value': '10.208.247.33'},
                    "root[0]['device_network']": {'new_value': 'tfnsw_network',
                                                  'old_value': 'rms_cameras_network'}}}'''
# above is the output of values_changed
try:
    diffData = difference['values_changed']
    keys = difference['values_changed'].keys()

    for item in keys:
        # if device_hostname exist in item
        if item.find('device_hostname') != -1:
            # item will be something like root[0]['device_hostname']
            hostkey = item
            hostkey_item = diffData[hostkey].get('new_value')
            old_hostkey_item = diffData[hostkey].get('old_value')

            if len(hostkey) > 0:
                # ['device_network'] may not exist in difference, so have to compare with original json file
                for i in range(0,len(data)):
                    if old_hostkey_item == data[i]['device_hostname']:
                        old_agency = data[i]['device_network']
                        orgi_sitename = data[i]['device_site']
                        temp_sitename = orgi_sitename.replace(":", ' ')
                        if old_agency == 'tnsw_network':
                            old_sitename = data[i]['device_hostname'][4:12]
                        else:
                            old_sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
                        old_directory = r"/home/cxu/pyats/{}".format(old_agency + '/' + old_sitename)
                for i in range(0, len(data1)):
                    host = data1[i]['device_hostname']
                    if hostkey_item == host:
                        agency = data1[i]['device_network']
                        orgi_sitename = data1[i]['device_site']
                        temp_sitename = orgi_sitename.replace(":", ' ')
                        if agency == 'tnsw_network':
                            sitename = data1[i]['device_hostname'][4:12]
                        else:
                            sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
                        directory = r"/home/cxu/pyats/{}".format(agency + '/' + sitename)
                        # if directory does not exist, create new directory, append info to newly created file.
                        # remove entry from old directory/file
                        # if directory already exist, remove old entry and add new entry in the same file.
                        if not os.path.exists(directory):
                            os.makedirs(directory)
                            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,
                                                                                                                 sitename)
                            shutil.copyfile(origin_file, target_file)
                            print('aaaaaaaa')
                            wb = load_workbook(target_file)
                            page = wb.active
                            info = [data1[i]['device_hostname'], data1[i]['device_native_ip'], 'c', '', 'ssh',
                                    data1[i]['device_type']]
                            page.append(info)
                            wb.save(filename=target_file)
                            if agency not in agencylist:
                                dic[agency] = []
                                dic[agency].append(sitename)
                            elif sitename not in dic[agency]:
                                dic[agency].append(sitename)
                            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(
                                old_agency,
                                old_sitename)
                            wb = load_workbook(target_file)
                            source = wb['Sheet1']
                            page = wb.active
                            hostlist = []
                            for cell in source['A']:
                                hostlist.append(cell.value)
                            if diffData[hostkey].get('old_value') in hostlist:
                                index = hostlist.index(diffData[item]['old_value']) + 1
                                # remove entry from hostname.xlsx since the device has been removed from json file
                                source.delete_rows(index)
                                wb.save(filename=target_file)
                                if old_agency not in agencylist:
                                    dic[old_agency] = []
                                    dic[old_agency].append(old_sitename)
                                elif old_sitename not in dic[old_agency]:
                                    dic[old_agency].append(old_sitename)
                        else:
                            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,
                                                                                                                 sitename)
                            wb = load_workbook(target_file)
                            source = wb['Sheet1']
                            page = wb.active
                            info = [data1[i]['device_hostname'], data1[i]['device_native_ip'], 'c', '', 'ssh',
                                    data1[i]['device_type']]
                            hostlist = []
                            for cell in source['A']:
                                hostlist.append(cell.value)
                            if diffData[hostkey].get('old_value') in hostlist:
                                index = hostlist.index(diffData[item]['old_value']) + 1
                                source.delete_rows(index)
                                page.append(info)
                                wb.save(filename=target_file)
                                if old_agency not in agencylist:
                                    dic[old_agency] = []
                                    dic[old_agency].append(old_sitename)
                                elif old_sitename not in dic[old_agency]:
                                    dic[old_agency].append(old_sitename)
                            elif diffData[hostkey].get('new_value') not in hostlist:
                                page.append(info)
                                wb.save(filename=target_file)
                                if agency not in agencylist:
                                    dic[agency] = []
                                    dic[agency].append(sitename)
                                elif sitename not in dic[agency]:
                                    dic[agency].append(sitename)
        # if only ip being changed
        elif item.find('device_native_ip') != -1:
            ipkey = item
            # get first part of key(root[0])
            index = ipkey[::-1][20:][::-1]
            index_str = index + '''['device_hostname']'''
            hostkey_exist = False
            for item in keys:
                if item == index_str:
                    hostkey_exist = True
                    break

            if len(ipkey) > 0 and hostkey_exist == False:
                for i in range(0, len(data1)):
                    ip = data1[i]['device_native_ip']
                    if diffData[ipkey]['new_value'] == ip:
                        agency = data1[i]['device_network']
                        orgi_sitename = data1[i]['device_site']
                        temp_sitename = orgi_sitename.replace(":", ' ')
                        if agency == 'tnsw_network':
                            sitename = data[i]['device_hostname'][4:12]
                        else:
                            sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
                        directory = r"/home/cxu/pyats/{}".format(agency + '/' + sitename)
                        if not os.path.exists(directory):
                            os.makedirs(directory)
                            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,
                                                                                                                 sitename)
                            shutil.copyfile(origin_file, target_file)
                            wb = load_workbook(target_file)
                            page = wb.active
                            info = [data1[i]['device_hostname'], data1[i]['device_native_ip'], 'c', '', 'ssh',
                                    data1[i]['device_type']]
                            page.append(info)
                            wb.save(filename=target_file)
                            if agency not in agencylist:
                                dic[agency] = []
                                dic[agency].append(sitename)
                            elif sitename not in dic[agency]:
                                dic[agency].append(sitename)
                        else:
                            target_file = r"/home/cxu/pyats/{}/{}/hostname.xlsx".format(agency,
                                                                                                                 sitename)
                            wb = load_workbook(target_file)
                            source = wb['Sheet1']
                            page = wb.active
                            info = [data1[i]['device_hostname'], data1[i]['device_native_ip'], 'c', '', 'ssh',
                                    data1[i]['device_type']]
                            iplist = []
                            for cell in source['B']:
                                iplist.append(cell.value)
                            if diffData[item]['old_value'] in iplist:
                                index = iplist.index(diffData[item]['old_value']) + 1
                                source.delete_rows(index)
                                page.append(info)
                                wb.save(filename=target_file)
                                if old_agency not in agencylist:
                                    dic[old_agency] = []
                                    dic[old_agency].append(old_sitename)
                                elif old_sitename not in dic[old_agency]:
                                    dic[old_agency].append(old_sitename)
                            else:
                                page.append(info)
                                wb.save(filename=target_file)
                                if agency not in agencylist:
                                    dic[agency] = []
                                    dic[agency].append(sitename)
                                elif sitename not in dic[agency]:
                                    dic[agency].append(sitename)
            else:
                pass

except KeyError:
    print('no change')
#
pprint(dic)
# pyats will throw error if there is other file(s) in the same folder as the hostname.xlsx
# to avoid the error, we will move the existing file(s) to tempfolder, create testbed file and move those file(s) back

tempfolder = "/home/cxu/pyats/tempfolder"
tempfile11 = "/home/cxu/pyats/tempfolder/file11.txt"
tempdevicetxt = "/home/cxu/pyats/tempfolder/device.txt"
# during comparing, new agency or site may appear, there is no agency list or device list text file in the new folder
# we will create the text files first(file11.txt,device.txt), which will be moved away and moved back later
for k,v in dic.items():
    for s in v:
        host_path = r"/home/cxu/pyats/{}/{}/".format(k,s)
        file11 = r"/home/cxu/pyats/{}/file11.txt".format(k)
        device_txt = r"/home/cxu/pyats/{}/{}/device.txt".format(k, s)
        if os.path.exists(file11):
            with open(file11, 'a+') as f:

                if s in f:
                    print('pass')
                else:
                    f.write(s + "\n")
                    with open(device_txt, 'w'): pass
                    with open(device_txt, 'a+') as file:
                        filepath = host_path + 'hostname.xlsx'
                        wb = load_workbook(filepath)
                        source = wb["Sheet1"]
                        for row in source.iter_rows(min_row=2, max_col=1, max_row=source.max_row):
                            for cell in row:
                                if cell.value in file:
                                    pass
                                else :
                                    file.write(cell.value + "\n")
        else:
            with open(file11, 'a+') as f:
                f.write(s + "\n")
                with open(device_txt, 'a+') as file:
                    filepath = host_path + 'hostname.xlsx'
                    wb = load_workbook(filepath)
                    source = wb["Sheet1"]
                    for row in source.iter_rows(min_row=2, max_col=1, max_row=source.max_row):
                        for cell in row:
                            if cell.value in file:
                                pass
                            else:
                                file.write(cell.value + "\n")

    for s in v:
        newdirectory = "/home/cxu/pyats/{}/{}/".format(k, s)
        t_file = r"/home/cxu/pyats/{}/{}/hostname.yaml".format(k , s)
        file11 = r"/home/cxu/pyats/{}/file11.txt".format(k)
        device_txt = r"/home/cxu/pyats/{}/{}/device.txt".format(k, s)
        if os.path.exists(t_file):
            os.remove(t_file)
            if os.path.exists(tempfile11):
                os.remove(tempfile11)
            if os.path.exists(tempdevicetxt):
                os.remove(tempdevicetxt)
            shutil.move(file11,tempfolder)
            shutil.move(device_txt,tempfolder)
            creator = File(path=newdirectory)
            creator.to_testbed_file(newdirectory)
            shutil.move(tempfile11,file11)
            shutil.move(tempdevicetxt,device_txt)
        else:
            if os.path.exists(tempfile11):
                os.remove(tempfile11)
            if os.path.exists(tempdevicetxt):
                os.remove(tempdevicetxt)
            shutil.move(file11, tempfolder)
            shutil.move(device_txt, tempfolder)
            creator = File(path=newdirectory)
            creator.to_testbed_file(newdirectory)
            shutil.move(tempfile11,file11)
            shutil.move(tempdevicetxt,device_txt)
