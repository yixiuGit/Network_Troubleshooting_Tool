import os
import json
import re
import shutil
from openpyxl import load_workbook, Workbook
from pyats.contrib.creators.file import File

with open('test.json') as f:
    data = json.load(f)

origin_file = r"/home/user/pyats/hostname.xlsx"
agency_list =[]
site_list = []
agency_site = {}
for i in range(0,len(data)):
    agency = data[i]['device_network']
    orgi_sitename = data[i]['device_site']
    temp_sitename = orgi_sitename.replace(":", ' ')
    if agency == 'tnsw_network' :
        sitename = data[i]['device_hostname'][4:12]
        # print(sitename)
    else:
        sitename = re.sub('[^A-Za-z0-9_ ]+', '', temp_sitename)
    if agency not in agency_site.keys():
        agency_site[agency] = []
        agency_site[agency].append(sitename)
        # print(agency_site[agency])
    elif sitename not in agency_site[agency]:
        agency_site[agency].append(sitename)
    directory = r"/home/user/pyats/{}".format(agency + '/' + sitename)
    if agency not in agency_list:
        agency_list.append(agency)
    if not os.path.exists(directory):
        os.makedirs(directory)
        target_file = r"/home/user/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
        shutil.copyfile(origin_file,target_file)
        wb = load_workbook(target_file)
        page = wb.active
        info = [data[i]['device_hostname'],data[i]['device_native_ip'],'c','cisco','ssh',data[i]['device_type']]
        page.append(info)
        wb.save(filename=target_file)
    else:
        target_file = "/home/user/pyats/{}/{}/hostname.xlsx".format(agency,sitename)
        wb = load_workbook(target_file)
        source = wb['Sheet1']
        page = wb.active
        info = [data[i]['device_hostname'], data[i]['device_native_ip'], 'c', 'cisco', 'ssh', data[i]['device_type']]
        hostlist = []
        for cell in source['A']:
            hostlist.append(cell.value)
        if data[i]['device_hostname'] in hostlist :
            print('eee')
        else:
            page.append(info)
            wb.save(filename=target_file)
# print(agency_list)
# print(agency_site)
for customer in agency_list:
    directory = "/home/user/pyats/{}".format(customer)
    creator = File(path=directory,recurse=True,encode_password=True)
    creator.to_testbed_file(directory)

for key, value in agency_site.items():
    agencydirectory = r"/home/user/pyats/{}/file11.txt".format(key)
    print(agencydirectory)
    with open(agencydirectory,'a') as file:
        for s in value:
            sitedirectory = r"/home/user/pyats/{}/{}/".format(key,s)
            devicedirectory = r"/home/user/pyats/{}/{}/device.txt".format(key,s)
            print(sitedirectory)
            file.write(s+"\n")
            with open (devicedirectory,'a') as f:
                filepath = sitedirectory + '/hostname.xlsx'
                wb = load_workbook(filepath)
                source = wb["Sheet1"]
                for row in source.iter_rows(min_row=2, max_col=1, max_row=source.max_row):
                    for cell in row:
                        f.write(cell.value+"\n")
